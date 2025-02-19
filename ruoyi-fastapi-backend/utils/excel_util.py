import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from typing import Dict, List


class ExcelUtil:
    """
    Excel操作类
    """

    @classmethod
    def __mapping_list(cls, list_data: List, mapping_dict: Dict):
        """
        工具方法：将list数据中的字段名映射为对应的中文字段名

        :param list_data: 数据列表
        :param mapping_dict: 映射字典
        :return: 映射后的数据列表
        """
        mapping_data = [{mapping_dict.get(key): item.get(key) for key in mapping_dict} for item in list_data]

        return mapping_data

    @classmethod
    def export_list2excel(cls, list_data: List, mapping_dict: Dict):
        """
        工具方法：将需要导出的list数据转化为对应excel的二进制数据

        :param list_data: 数据列表
        :param mapping_dict: 映射字典
        :return: list数据对应excel的二进制数据
        """
        mapping_data = cls.__mapping_list(list_data, mapping_dict)
        df = pd.DataFrame(mapping_data)
        binary_data = io.BytesIO()
        df.to_excel(binary_data, index=False, engine='openpyxl')
        binary_data = binary_data.getvalue()

        return binary_data

    @classmethod
    def get_excel_template(cls, header_list: List, selector_header_list: List, option_list: List[Dict]):
        """
        工具方法：将需要导出的list数据转化为对应excel的二进制数据

        :param header_list: 表头数据列表
        :param selector_header_list: 需要设置为选择器格式的表头数据列表
        :param option_list: 选择器格式的表头预设的选项列表
        :return: 模板excel的二进制数据
        """
        # 创建Excel工作簿
        wb = Workbook()
        # 选择默认的活动工作表
        ws = wb.active

        # 设置表头文字
        headers = header_list

        # 设置表头背景样式为灰色，前景色为白色
        header_fill = PatternFill(start_color='ababab', end_color='ababab', fill_type='solid')

        # 将表头写入第一行
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            # 设置列宽度为16
            ws.column_dimensions[chr(64 + col_num)].width = 12
            # 设置水平居中对齐
            cell.alignment = Alignment(horizontal='center')

        # 设置选择器的预设选项
        options = option_list

        # 获取selector_header的字母索引
        for selector_header in selector_header_list:
            column_selector_header_index = headers.index(selector_header) + 1

            # 创建数据有效性规则
            header_option = []
            for option in options:
                if option.get(selector_header):
                    header_option = option.get(selector_header)
            dv = DataValidation(type='list', formula1=f'"{",".join(header_option)}"')
            # 设置数据有效性规则的起始单元格和结束单元格
            dv.add(
                f'{get_column_letter(column_selector_header_index)}2:{get_column_letter(column_selector_header_index)}1048576'
            )
            # 添加数据有效性规则到工作表
            ws.add_data_validation(dv)

        # 保存Excel文件为字节类型的数据
        file = io.BytesIO()
        wb.save(file)
        file.seek(0)

        # 读取字节数据
        excel_data = file.getvalue()

        return excel_data
