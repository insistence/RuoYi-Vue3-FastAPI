import io
import os
import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.engine.row import Row
from sqlalchemy.orm.collections import InstrumentedList
from typing import Any, Dict, List, Literal, Union
from config.database import Base
from config.env import CachePathConfig


def worship():
    print("""
////////////////////////////////////////////////////////////////////
//                          _ooOoo_                               //
//                         o8888888o                              //
//                         88" . "88                              //
//                         (| ^_^ |)                              //
//                         O\  =  /O                              //
//                      ____/`---'\____                           //
//                    .'  \\|     |//  `.                         //
//                   /  \\|||  :  |||//  \                        //
//                  /  _||||| -:- |||||-  \                       //
//                  |   | \\\  -  /// |   |                       //
//                  | \_|  ''\---/''  |   |                       //
//                  \  .-\__  `-`  ___/-. /                       //
//                ___`. .'  /--.--\  `. . ___                     //
//              ."" '<  `.___\_<|>_/___.'  >'"".                  //
//            | | :  `- \`.;`\ _ /`;.`/ - ` : | |                 //
//            \  \ `-.   \_ __\ /__ _/   .-` /  /                 //
//      ========`-.____`-.___\_____/___.-`____.-'========         //
//                           `=---='                              //
//      ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^        //
//             佛祖保佑       永不宕机      永无BUG                 //
////////////////////////////////////////////////////////////////////
    """)


class SqlalchemyUtil:
    """
    sqlalchemy工具类
    """

    @classmethod
    def base_to_dict(
        cls, obj: Union[Base, Dict], transform_case: Literal['no_case', 'snake_to_camel', 'camel_to_snake'] = 'no_case'
    ):
        """
        将sqlalchemy模型对象转换为字典

        :param obj: sqlalchemy模型对象或普通字典
        :param transform_case: 转换得到的结果形式，可选的有'no_case'(不转换)、'snake_to_camel'(下划线转小驼峰)、'camel_to_snake'(小驼峰转下划线)，默认为'no_case'
        :return: 字典结果
        """
        if isinstance(obj, Base):
            base_dict = obj.__dict__.copy()
            base_dict.pop('_sa_instance_state', None)
            for name, value in base_dict.items():
                if isinstance(value, InstrumentedList):
                    base_dict[name] = cls.serialize_result(value, 'snake_to_camel')
        elif isinstance(obj, dict):
            base_dict = obj.copy()
        if transform_case == 'snake_to_camel':
            return {CamelCaseUtil.snake_to_camel(k): v for k, v in base_dict.items()}
        elif transform_case == 'camel_to_snake':
            return {SnakeCaseUtil.camel_to_snake(k): v for k, v in base_dict.items()}

        return base_dict

    @classmethod
    def serialize_result(
        cls, result: Any, transform_case: Literal['no_case', 'snake_to_camel', 'camel_to_snake'] = 'no_case'
    ):
        """
        将sqlalchemy查询结果序列化

        :param result: sqlalchemy查询结果
        :param transform_case: 转换得到的结果形式，可选的有'no_case'(不转换)、'snake_to_camel'(下划线转小驼峰)、'camel_to_snake'(小驼峰转下划线)，默认为'no_case'
        :return: 序列化结果
        """
        if isinstance(result, (Base, dict)):
            return cls.base_to_dict(result, transform_case)
        elif isinstance(result, list):
            return [cls.serialize_result(row, transform_case) for row in result]
        elif isinstance(result, Row):
            if all([isinstance(row, Base) for row in result]):
                return [cls.base_to_dict(row, transform_case) for row in result]
            elif any([isinstance(row, Base) for row in result]):
                return [cls.serialize_result(row, transform_case) for row in result]
            else:
                result_dict = result._asdict()
                if transform_case == 'snake_to_camel':
                    return {CamelCaseUtil.snake_to_camel(k): v for k, v in result_dict.items()}
                elif transform_case == 'camel_to_snake':
                    return {SnakeCaseUtil.camel_to_snake(k): v for k, v in result_dict.items()}
                return result_dict
        return result


class CamelCaseUtil:
    """
    下划线形式(snake_case)转小驼峰形式(camelCase)工具方法
    """

    @classmethod
    def snake_to_camel(cls, snake_str: str):
        """
        下划线形式字符串(snake_case)转换为小驼峰形式字符串(camelCase)

        :param snake_str: 下划线形式字符串
        :return: 小驼峰形式字符串
        """
        # 分割字符串
        words = snake_str.split('_')
        # 小驼峰命名，第一个词首字母小写，其余词首字母大写
        return words[0] + ''.join(word.capitalize() for word in words[1:])

    @classmethod
    def transform_result(cls, result: Any):
        """
        针对不同类型将下划线形式(snake_case)批量转换为小驼峰形式(camelCase)方法

        :param result: 输入数据
        :return: 小驼峰形式结果
        """
        return SqlalchemyUtil.serialize_result(result=result, transform_case='snake_to_camel')


class SnakeCaseUtil:
    """
    小驼峰形式(camelCase)转下划线形式(snake_case)工具方法
    """

    @classmethod
    def camel_to_snake(cls, camel_str: str):
        """
        小驼峰形式字符串(camelCase)转换为下划线形式字符串(snake_case)

        :param camel_str: 小驼峰形式字符串
        :return: 下划线形式字符串
        """
        # 在大写字母前添加一个下划线，然后将整个字符串转为小写
        words = re.sub('(.)([A-Z][a-z]+)', r'\1_\2', camel_str)
        return re.sub('([a-z0-9])([A-Z])', r'\1_\2', words).lower()

    @classmethod
    def transform_result(cls, result: Any):
        """
        针对不同类型将下划线形式(snake_case)批量转换为小驼峰形式(camelCase)方法

        :param result: 输入数据
        :return: 小驼峰形式结果
        """
        return SqlalchemyUtil.serialize_result(result=result, transform_case='camel_to_snake')


def bytes2human(n, format_str='%(value).1f%(symbol)s'):
    """Used by various scripts. See:
    http://goo.gl/zeJZl

    >>> bytes2human(10000)
    '9.8K'
    >>> bytes2human(100001221)
    '95.4M'
    """
    symbols = ('B', 'KB', 'MB', 'GB', 'TB', 'PB', 'EB', 'ZB', 'YB')
    prefix = {}
    for i, s in enumerate(symbols[1:]):
        prefix[s] = 1 << (i + 1) * 10
    for symbol in reversed(symbols[1:]):
        if n >= prefix[symbol]:
            value = float(n) / prefix[symbol]
            return format_str % locals()
    return format_str % dict(symbol=symbols[0], value=n)


def bytes2file_response(bytes_info):
    yield bytes_info


def export_list2excel(list_data: List):
    """
    工具方法：将需要导出的list数据转化为对应excel的二进制数据

    :param list_data: 数据列表
    :return: 字典信息对应excel的二进制数据
    """
    df = pd.DataFrame(list_data)
    binary_data = io.BytesIO()
    df.to_excel(binary_data, index=False, engine='openpyxl')
    binary_data = binary_data.getvalue()

    return binary_data


def get_excel_template(header_list: List, selector_header_list: List, option_list: List[dict]):
    """
    工具方法：将需要导出的list数据转化为对应excel的二进制数据

    :param header_list: 表头数据列表
    :param selector_header_list: 需要设置为选择器格式的表头数据列表
    :param option_list: 选择器格式的表头预设的选项列表
    :return: 模板excel的二进制数据
    """
    # 创建Excel工作簿
    wb = Workbook()
    # 选择默认的活动工作表
    ws = wb.active

    # 设置表头文字
    headers = header_list

    # 设置表头背景样式为灰色，前景色为白色
    header_fill = PatternFill(start_color='ababab', end_color='ababab', fill_type='solid')

    # 将表头写入第一行
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        # 设置列宽度为16
        ws.column_dimensions[chr(64 + col_num)].width = 12
        # 设置水平居中对齐
        cell.alignment = Alignment(horizontal='center')

    # 设置选择器的预设选项
    options = option_list

    # 获取selector_header的字母索引
    for selector_header in selector_header_list:
        column_selector_header_index = headers.index(selector_header) + 1

        # 创建数据有效性规则
        header_option = []
        for option in options:
            if option.get(selector_header):
                header_option = option.get(selector_header)
        dv = DataValidation(type='list', formula1=f'"{",".join(header_option)}"')
        # 设置数据有效性规则的起始单元格和结束单元格
        dv.add(
            f'{get_column_letter(column_selector_header_index)}2:{get_column_letter(column_selector_header_index)}1048576'
        )
        # 添加数据有效性规则到工作表
        ws.add_data_validation(dv)

    # 保存Excel文件为字节类型的数据
    file = io.BytesIO()
    wb.save(file)
    file.seek(0)

    # 读取字节数据
    excel_data = file.getvalue()

    return excel_data


def get_filepath_from_url(url: str):
    """
    工具方法：根据请求参数获取文件路径

    :param url: 请求参数中的url参数
    :return: 文件路径
    """
    file_info = url.split('?')[1].split('&')
    task_id = file_info[0].split('=')[1]
    file_name = file_info[1].split('=')[1]
    task_path = file_info[2].split('=')[1]
    filepath = os.path.join(CachePathConfig.PATH, task_path, task_id, file_name)

    return filepath
